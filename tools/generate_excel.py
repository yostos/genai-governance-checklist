#!/usr/bin/env python3
"""
生成AI利用ガイドライン チェックリスト — Excel 生成スクリプト

src/ch01-governance.md – src/ch07-document-quality.md をパースし、
セルフチェック用 Excel ワークシートを生成する。

Usage:
    python tools/generate_excel.py

Output:
    excel/genai-governance-checklist.xlsx

仕様書: tools/docs/excel-spec.md
"""

import re
import sys
from pathlib import Path

from openpyxl import Workbook
from openpyxl.formatting.rule import CellIsRule
from openpyxl.styles import (
    Alignment,
    Border,
    Font,
    PatternFill,
    Protection,
    Side,
)
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.datavalidation import DataValidation
from openpyxl.worksheet.page import PageMargins
from openpyxl.worksheet.properties import PageSetupProperties

# ---------------------------------------------------------------------------
# Configuration
# ---------------------------------------------------------------------------

REPO_ROOT = Path(__file__).resolve().parent.parent
SRC_DIR = REPO_ROOT / "src"
OUTPUT_DIR = REPO_ROOT / "excel"
OUTPUT_FILE = OUTPUT_DIR / "genai-governance-checklist.xlsx"

CHAPTER_FILES = [
    "ch01-governance.md",
    "ch02-risk-mapping.md",
    "ch03-input-data.md",
    "ch04-output-management.md",
    "ch05-trustworthiness.md",
    "ch06-incident-response.md",
    "ch07-document-quality.md",
]

REFERENCE_TAGS = ["NIST", "NIST-GAI", "METI", "JDLA", "IPA", "FUJITSU", "EU-AIA"]

CHAPTER_COLORS = {
    1: "E2EFDA",  # 薄緑
    2: "DEEBF7",  # 薄青
    3: "FFF2CC",  # 薄黄
    4: "F2DCDB",  # 薄紅
    5: "E2D9F3",  # 薄紫
    6: "FCE4D6",  # 薄橙
    7: "D6DCE4",  # 薄灰
}

CHECK_STATUSES = ["対応済", "一部対応", "未対応", "該当なし"]

FONT_NAME = "メイリオ"

# Colors
DARK_NAVY = "1F3864"
DARK_BLUE_GRAY = "44546A"
LIGHT_GRAY = "D9E2F3"
WHITE = "FFFFFF"

HEADER_ROW = 3
DATA_START_ROW = 4

# Column definitions: (header_name, width_in_chars)
COLUMNS = [
    ("チェック結果", 12),     # A
    ("項目番号", 8),           # B
    ("章", 5),                 # C
    ("章タイトル", 28),        # D
    ("節", 6),                 # E
    ("節タイトル", 24),        # F
    ("準拠レベル", 14),        # G
    ("チェック項目", 60),      # H
    ("NIST", 6),               # I
    ("NIST-GAI", 9),           # J
    ("METI", 6),               # K
    ("JDLA", 6),               # L
    ("IPA", 5),                # M
    ("FUJITSU", 9),            # N
    ("EU-AIA", 8),             # O
    ("備考", 30),              # P
    ("対応状況メモ", 40),      # Q
]

# ---------------------------------------------------------------------------
# Regex patterns
# ---------------------------------------------------------------------------

RE_CHAPTER = re.compile(r"^# (\d+)\.\s+(.+)$")
RE_SECTION = re.compile(r"^## (\d+\.\d+)\s+(.+)$")
RE_ITEM = re.compile(
    r"^- (\d+\.\d+\.[A-Z])\.\s+\[(Required|Recommended|Option)\]\s+(.+)$"
)
RE_TAG = re.compile(
    r"\[(NIST(?:-GAI)?|METI|JDLA|IPA|FUJITSU|EU-AIA)(?::[^\]]+)?\]"
)

# ---------------------------------------------------------------------------
# Parsing
# ---------------------------------------------------------------------------


def parse_chapters():
    """Parse all chapter markdown files.

    Returns:
        items: list of dicts with checklist item data
        chapters: dict mapping chapter_num -> chapter_title
    """
    items = []
    chapters = {}

    for filename in CHAPTER_FILES:
        filepath = SRC_DIR / filename
        if not filepath.exists():
            print(f"Warning: {filepath} not found, skipping.", file=sys.stderr)
            continue

        ch_num = None
        ch_title = None
        sec_num = None
        sec_title = None

        for line in filepath.read_text(encoding="utf-8").splitlines():
            m = RE_CHAPTER.match(line)
            if m:
                ch_num = int(m.group(1))
                ch_title = m.group(2).strip()
                chapters[ch_num] = ch_title
                continue

            m = RE_SECTION.match(line)
            if m:
                sec_num = m.group(1)
                sec_title = m.group(2).strip()
                continue

            m = RE_ITEM.match(line)
            if m:
                number = m.group(1)
                level = m.group(2)
                raw_text = m.group(3).strip()

                tags = set(RE_TAG.findall(raw_text))
                clean_text = RE_TAG.sub("", raw_text).strip()

                items.append(
                    {
                        "number": number,
                        "chapter_num": ch_num,
                        "chapter_title": ch_title,
                        "section_num": sec_num,
                        "section_title": sec_title,
                        "level": level,
                        "text": clean_text,
                        "tags": tags,
                    }
                )

    return items, chapters


# ---------------------------------------------------------------------------
# Excel building: Checklist sheet
# ---------------------------------------------------------------------------


def build_checklist_sheet(wb, items):
    """Build the main checklist sheet. Returns data_end_row."""
    ws = wb.active
    ws.title = "チェックリスト"

    num_items = len(items)
    data_end_row = DATA_START_ROW + num_items - 1
    last_col = len(COLUMNS)
    last_col_letter = get_column_letter(last_col)

    # --- Column widths ---
    for i, (_name, width) in enumerate(COLUMNS, 1):
        ws.column_dimensions[get_column_letter(i)].width = width

    # --- Title row (row 1) ---
    ws.merge_cells(f"A1:{last_col_letter}1")
    cell = ws.cell(
        row=1, column=1, value="生成AI利用ガイドライン チェックリスト"
    )
    cell.font = Font(name=FONT_NAME, size=16, bold=True, color=WHITE)
    cell.fill = PatternFill(
        start_color=DARK_NAVY, end_color=DARK_NAVY, fill_type="solid"
    )
    cell.alignment = Alignment(horizontal="center", vertical="center")
    ws.row_dimensions[1].height = 36

    # --- Input row (row 2) ---
    ws.merge_cells(f"A2:{last_col_letter}2")
    cell = ws.cell(
        row=2,
        column=1,
        value="組織名: ＿＿＿＿＿　記入者: ＿＿＿＿＿　記入日: ＿＿＿＿年＿＿月＿＿日",
    )
    cell.font = Font(name=FONT_NAME, size=10)
    cell.fill = PatternFill(
        start_color=LIGHT_GRAY, end_color=LIGHT_GRAY, fill_type="solid"
    )
    cell.alignment = Alignment(vertical="center")
    ws.row_dimensions[2].height = 28

    # --- Header row (row 3) ---
    header_font = Font(name=FONT_NAME, size=11, bold=True, color=WHITE)
    header_fill = PatternFill(
        start_color=DARK_NAVY, end_color=DARK_NAVY, fill_type="solid"
    )
    header_align = Alignment(
        horizontal="center", vertical="center", wrap_text=True
    )
    white_border = Border(
        left=Side(style="thin", color=WHITE),
        right=Side(style="thin", color=WHITE),
        top=Side(style="thin", color=WHITE),
        bottom=Side(style="thin", color=WHITE),
    )

    for i, (name, _) in enumerate(COLUMNS, 1):
        cell = ws.cell(row=HEADER_ROW, column=i, value=name)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = header_align
        cell.border = white_border
    ws.row_dimensions[HEADER_ROW].height = 40

    # --- Data rows ---
    data_font = Font(name=FONT_NAME, size=10)
    wrap_top = Alignment(vertical="top", wrap_text=True)
    center = Alignment(horizontal="center", vertical="center")

    for idx, item in enumerate(items):
        row = DATA_START_ROW + idx
        ch_num = item["chapter_num"]
        ch_fill = PatternFill(
            start_color=CHAPTER_COLORS.get(ch_num, "FFFFFF"),
            end_color=CHAPTER_COLORS.get(ch_num, "FFFFFF"),
            fill_type="solid",
        )

        row_data = [
            "",                         # A: チェック結果
            item["number"],             # B: 項目番号
            ch_num,                     # C: 章
            item["chapter_title"],      # D: 章タイトル
            item["section_num"],        # E: 節
            item["section_title"],      # F: 節タイトル
            item["level"],              # G: 準拠レベル
            item["text"],               # H: チェック項目
        ]
        # Reference tags (I–O)
        for tag in REFERENCE_TAGS:
            row_data.append("○" if tag in item["tags"] else "")
        # User columns (P–Q)
        row_data.append("")  # P: 備考
        row_data.append("")  # Q: 対応状況メモ

        for col_idx, value in enumerate(row_data, 1):
            cell = ws.cell(row=row, column=col_idx, value=value)
            cell.font = data_font

            # Chapter background for B–O (columns 2–15)
            if 2 <= col_idx <= 15:
                cell.fill = ch_fill

            # Text wrapping for H, P, Q
            if col_idx in (8, 16, 17):
                cell.alignment = wrap_top

            # Center alignment for reference tag columns I–O (9–15)
            if 9 <= col_idx <= 15:
                cell.alignment = center

    # --- Data validation: A column dropdown ---
    dv = DataValidation(
        type="list",
        formula1='"対応済,一部対応,未対応,該当なし"',
        allow_blank=True,
    )
    dv.error = "対応済・一部対応・未対応・該当なしから選択してください"
    dv.errorTitle = "入力エラー"
    dv.showErrorMessage = True
    ws.add_data_validation(dv)
    dv.add(f"A{DATA_START_ROW}:A{data_end_row}")

    # --- Conditional formatting ---
    a_range = f"A{DATA_START_ROW}:A{data_end_row}"
    g_range = f"G{DATA_START_ROW}:G{data_end_row}"
    tag_range = f"I{DATA_START_ROW}:O{data_end_row}"

    # A column: check result
    for status, bg, fg, italic in [
        ("対応済", "C6EFCE", "006100", False),
        ("一部対応", "FFEB9C", "9C6500", False),
        ("未対応", "FFC7CE", "9C0006", False),
        ("該当なし", "D9D9D9", "808080", True),
    ]:
        ws.conditional_formatting.add(
            a_range,
            CellIsRule(
                operator="equal",
                formula=[f'"{status}"'],
                fill=PatternFill(
                    start_color=bg, end_color=bg, fill_type="solid"
                ),
                font=Font(name=FONT_NAME, color=fg, italic=italic),
            ),
        )

    # G column: enforcement level
    for level, bg, fg, bold in [
        ("Required", "F4B084", "843C0C", True),
        ("Recommended", "BDD7EE", "1F4E79", False),
        ("Option", "E2EFDA", "375623", False),
    ]:
        ws.conditional_formatting.add(
            g_range,
            CellIsRule(
                operator="equal",
                formula=[f'"{level}"'],
                fill=PatternFill(
                    start_color=bg, end_color=bg, fill_type="solid"
                ),
                font=Font(name=FONT_NAME, color=fg, bold=bold),
            ),
        )

    # I–O columns: reference tag highlight
    ws.conditional_formatting.add(
        tag_range,
        CellIsRule(
            operator="equal",
            formula=['"○"'],
            fill=PatternFill(
                start_color="D6E4F0", end_color="D6E4F0", fill_type="solid"
            ),
        ),
    )

    # --- Auto filter ---
    ws.auto_filter.ref = f"A{HEADER_ROW}:{last_col_letter}{data_end_row}"

    # --- Freeze panes ---
    ws.freeze_panes = "B4"

    # --- Sheet protection (no password) ---
    ws.protection.sheet = True
    # Unlock editable columns: A (1), P (16), Q (17)
    for r in range(DATA_START_ROW, data_end_row + 1):
        for c in (1, 16, 17):
            ws.cell(row=r, column=c).protection = Protection(locked=False)
    # Unlock input row
    ws.cell(row=2, column=1).protection = Protection(locked=False)

    # --- Print settings ---
    ws.sheet_properties.pageSetUpPr = PageSetupProperties(fitToPage=True)
    ws.page_setup.paperSize = 8  # A3
    ws.page_setup.orientation = "landscape"
    ws.page_setup.fitToWidth = 1
    ws.page_setup.fitToHeight = 0
    ws.page_margins = PageMargins(
        left=0.59, right=0.59, top=0.59, bottom=0.59,
        header=0.3, footer=0.3,
    )
    ws.print_title_rows = "1:3"
    ws.print_title_cols = "A:B"
    ws.oddHeader.left.text = "生成AI利用ガイドライン チェックリスト"
    ws.oddHeader.right.text = "&P"
    ws.oddFooter.center.text = "Confidential"

    return data_end_row


# ---------------------------------------------------------------------------
# Excel building: Summary sheet
# ---------------------------------------------------------------------------


def build_summary_sheet(wb, chapters, data_end_row):
    """Build the summary dashboard sheet."""
    ws = wb.create_sheet("サマリー")

    # Sheet reference for formulas
    cl = "'チェックリスト'"
    a_ref = f"{cl}!$A${DATA_START_ROW}:$A${data_end_row}"
    g_ref = f"{cl}!$G${DATA_START_ROW}:$G${data_end_row}"
    c_ref = f"{cl}!$C${DATA_START_ROW}:$C${data_end_row}"

    total_items = data_end_row - DATA_START_ROW + 1

    # Common styles
    title_font = Font(name=FONT_NAME, size=14, bold=True, color=WHITE)
    title_fill = PatternFill(
        start_color=DARK_NAVY, end_color=DARK_NAVY, fill_type="solid"
    )
    hdr_font = Font(name=FONT_NAME, size=10, bold=True, color=WHITE)
    hdr_fill = PatternFill(
        start_color=DARK_BLUE_GRAY, end_color=DARK_BLUE_GRAY, fill_type="solid"
    )
    data_font = Font(name=FONT_NAME, size=10)
    bold_font = Font(name=FONT_NAME, size=10, bold=True)
    ctr = Alignment(horizontal="center", vertical="center")
    pct_fmt = "0.0%"

    # Column widths
    for col_letter, w in [
        ("A", 22), ("B", 10), ("C", 10), ("D", 10),
        ("E", 10), ("F", 10), ("G", 10), ("H", 12), ("I", 12),
    ]:
        ws.column_dimensions[col_letter].width = w

    # -------------------------------------------------------------------
    # Section 1: Overall Summary (rows 1–8)
    # -------------------------------------------------------------------
    ws.merge_cells("A1:C1")
    cell = ws.cell(row=1, column=1, value="チェック結果 全体サマリー")
    cell.font = title_font
    cell.fill = title_fill
    cell.alignment = ctr
    ws.row_dimensions[1].height = 32

    for col, text in enumerate(["ステータス", "件数", "割合"], 1):
        cell = ws.cell(row=2, column=col, value=text)
        cell.font = hdr_font
        cell.fill = hdr_fill
        cell.alignment = ctr

    statuses_with_formula = [
        ("対応済", f'=COUNTIF({a_ref},"対応済")'),
        ("一部対応", f'=COUNTIF({a_ref},"一部対応")'),
        ("未対応", f'=COUNTIF({a_ref},"未対応")'),
        ("該当なし", f'=COUNTIF({a_ref},"該当なし")'),
        ("未記入", f"=COUNTBLANK({a_ref})"),
    ]
    for i, (label, formula) in enumerate(statuses_with_formula):
        row = 3 + i
        ws.cell(row=row, column=1, value=label).font = data_font
        ws.cell(row=row, column=2).value = formula
        ws.cell(row=row, column=2).font = data_font
        ws.cell(row=row, column=2).alignment = ctr
        ws.cell(row=row, column=3).value = f"=B{row}/{total_items}"
        ws.cell(row=row, column=3).font = data_font
        ws.cell(row=row, column=3).number_format = pct_fmt
        ws.cell(row=row, column=3).alignment = ctr

    ws.cell(row=8, column=1, value="合計").font = bold_font
    ws.cell(row=8, column=2, value=total_items).font = bold_font
    ws.cell(row=8, column=2).alignment = ctr

    # -------------------------------------------------------------------
    # Section 2: Level Summary (rows 10–14)
    # -------------------------------------------------------------------
    ws.merge_cells("A10:H10")
    cell = ws.cell(row=10, column=1, value="準拠レベル別 チェック結果")
    cell.font = title_font
    cell.fill = title_fill
    cell.alignment = ctr
    ws.row_dimensions[10].height = 32

    lv_headers = [
        "レベル", "項目数", "対応済", "一部対応",
        "未対応", "該当なし", "未記入", "対応率",
    ]
    for col, text in enumerate(lv_headers, 1):
        cell = ws.cell(row=11, column=col, value=text)
        cell.font = hdr_font
        cell.fill = hdr_fill
        cell.alignment = ctr

    for i, level in enumerate(["Required", "Recommended", "Option"]):
        row = 12 + i
        ws.cell(row=row, column=1, value=level).font = data_font
        # Item count
        ws.cell(row=row, column=2).value = f'=COUNTIF({g_ref},"{level}")'
        ws.cell(row=row, column=2).font = data_font
        ws.cell(row=row, column=2).alignment = ctr
        # Status counts
        for j, status in enumerate(CHECK_STATUSES):
            ws.cell(row=row, column=3 + j).value = (
                f'=COUNTIFS({g_ref},"{level}",{a_ref},"{status}")'
            )
            ws.cell(row=row, column=3 + j).font = data_font
            ws.cell(row=row, column=3 + j).alignment = ctr
        # 未記入 = total - sum of statuses
        ws.cell(row=row, column=7).value = f"=B{row}-SUM(C{row}:F{row})"
        ws.cell(row=row, column=7).font = data_font
        ws.cell(row=row, column=7).alignment = ctr
        # 対応率
        ws.cell(row=row, column=8).value = f"=IFERROR(C{row}/B{row},0)"
        ws.cell(row=row, column=8).font = data_font
        ws.cell(row=row, column=8).number_format = pct_fmt
        ws.cell(row=row, column=8).alignment = ctr

    # -------------------------------------------------------------------
    # Section 3: Chapter Summary (rows 16–24)
    # -------------------------------------------------------------------
    ws.merge_cells("A16:I16")
    cell = ws.cell(row=16, column=1, value="章別 チェック結果")
    cell.font = title_font
    cell.fill = title_fill
    cell.alignment = ctr
    ws.row_dimensions[16].height = 32

    ch_headers = [
        "章", "項目数", "対応済", "一部対応",
        "未対応", "該当なし", "未記入", "対応率", "必須対応率",
    ]
    for col, text in enumerate(ch_headers, 1):
        cell = ws.cell(row=17, column=col, value=text)
        cell.font = hdr_font
        cell.fill = hdr_fill
        cell.alignment = ctr

    for ch_num in sorted(chapters.keys()):
        row = 17 + ch_num  # ch1 → row 18, ch7 → row 24
        ch_label = f"{ch_num}. {chapters[ch_num]}"
        ws.cell(row=row, column=1, value=ch_label).font = data_font
        # Item count
        ws.cell(row=row, column=2).value = f"=COUNTIF({c_ref},{ch_num})"
        ws.cell(row=row, column=2).font = data_font
        ws.cell(row=row, column=2).alignment = ctr
        # Status counts
        for j, status in enumerate(CHECK_STATUSES):
            ws.cell(row=row, column=3 + j).value = (
                f'=COUNTIFS({c_ref},{ch_num},{a_ref},"{status}")'
            )
            ws.cell(row=row, column=3 + j).font = data_font
            ws.cell(row=row, column=3 + j).alignment = ctr
        # 未記入
        ws.cell(row=row, column=7).value = f"=B{row}-SUM(C{row}:F{row})"
        ws.cell(row=row, column=7).font = data_font
        ws.cell(row=row, column=7).alignment = ctr
        # 対応率
        ws.cell(row=row, column=8).value = f"=IFERROR(C{row}/B{row},0)"
        ws.cell(row=row, column=8).font = data_font
        ws.cell(row=row, column=8).number_format = pct_fmt
        ws.cell(row=row, column=8).alignment = ctr
        # 必須対応率
        req_total = f'COUNTIFS({c_ref},{ch_num},{g_ref},"Required")'
        req_done = (
            f'COUNTIFS({c_ref},{ch_num},{g_ref},"Required",'
            f'{a_ref},"対応済")'
        )
        ws.cell(row=row, column=9).value = f"=IFERROR({req_done}/{req_total},0)"
        ws.cell(row=row, column=9).font = data_font
        ws.cell(row=row, column=9).number_format = pct_fmt
        ws.cell(row=row, column=9).alignment = ctr

    # -------------------------------------------------------------------
    # Conditional formatting on percentage cells
    # -------------------------------------------------------------------
    green = PatternFill(start_color="C6EFCE", end_color="C6EFCE", fill_type="solid")
    yellow = PatternFill(start_color="FFEB9C", end_color="FFEB9C", fill_type="solid")
    red = PatternFill(start_color="FFC7CE", end_color="FFC7CE", fill_type="solid")

    pct_ranges = ["C3:C7", "H12:H14", "H18:I24"]
    for rng in pct_ranges:
        ws.conditional_formatting.add(
            rng,
            CellIsRule(
                operator="greaterThanOrEqual",
                formula=["1"],
                fill=green,
                stopIfTrue=True,
            ),
        )
        ws.conditional_formatting.add(
            rng,
            CellIsRule(
                operator="greaterThanOrEqual",
                formula=["0.8"],
                fill=yellow,
                stopIfTrue=True,
            ),
        )
        ws.conditional_formatting.add(
            rng,
            CellIsRule(
                operator="lessThan",
                formula=["0.8"],
                fill=red,
                stopIfTrue=True,
            ),
        )

    # -------------------------------------------------------------------
    # Section 4: Guide text (rows 26+)
    # -------------------------------------------------------------------
    guide_start = 26
    ws.cell(row=guide_start, column=1, value="チェック結果の入力方法").font = (
        Font(name=FONT_NAME, size=12, bold=True)
    )

    guide_lines = [
        "",
        "チェック結果の選択基準:",
        "　対応済:　　ガイドラインに該当する内容が記載されている",
        "　一部対応:　内容は含まれているが不十分、または表現が曖昧",
        "　未対応:　　ガイドラインに該当する記載がない",
        "　該当なし:　自組織の状況に照らして対応不要と判断した項目",
        "",
        "推奨する進め方:",
        "　1. まず「Required」項目だけをフィルタで表示し、すべてチェックする",
        "　2. 次に「Recommended」項目を確認する",
        "　3. 最後に「Option」項目を組織の状況に応じて検討する",
        "　4.「該当なし」とした項目は、備考欄にその理由を記入する",
    ]
    for i, text in enumerate(guide_lines):
        ws.cell(
            row=guide_start + 1 + i, column=1, value=text
        ).font = data_font

    # --- Sheet protection ---
    ws.protection.sheet = True

    # --- Print settings ---
    ws.page_setup.paperSize = 9  # A4
    ws.page_setup.orientation = "portrait"
    ws.sheet_properties.pageSetUpPr = PageSetupProperties(fitToPage=True)
    ws.page_setup.fitToWidth = 1
    ws.page_setup.fitToHeight = 1
    ws.page_margins = PageMargins(
        left=0.79, right=0.79, top=0.79, bottom=0.79,
        header=0.3, footer=0.3,
    )


# ---------------------------------------------------------------------------
# Main
# ---------------------------------------------------------------------------


def main():
    items, chapters = parse_chapters()
    print(f"Parsed {len(items)} checklist items from {len(chapters)} chapters.")

    if not items:
        print("Error: No items found. Check source files.", file=sys.stderr)
        sys.exit(1)

    OUTPUT_DIR.mkdir(parents=True, exist_ok=True)

    wb = Workbook()
    data_end_row = build_checklist_sheet(wb, items)
    build_summary_sheet(wb, chapters, data_end_row)

    wb.save(OUTPUT_FILE)
    print(f"Generated: {OUTPUT_FILE}")


if __name__ == "__main__":
    main()
